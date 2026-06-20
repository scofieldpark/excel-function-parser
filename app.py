import io
import json
import openpyxl
from flask import Flask, request, jsonify, render_template

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB


def cell_value_str(value):
    if value is None:
        return ''
    from datetime import datetime, date
    if isinstance(value, (datetime, date)):
        return str(value)
    return str(value)


def parse_workbook(file_bytes):
    """
    openpyxl로 같은 파일을 두 번 읽음:
      - data_only=False : 수식(formula) 원본
      - data_only=True  : 캐시된 계산값
    """
    wb_formula = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    wb_value   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    sheets = {}
    for name in wb_formula.sheetnames:
        ws_f = wb_formula[name]
        ws_v = wb_value[name]

        rows = []
        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
            row_cells = []
            for cf, cv in zip(row_f, row_v):
                raw_formula = cf.value
                raw_value   = cv.value

                formula_str = cell_value_str(raw_formula)
                value_str   = cell_value_str(raw_value)
                is_formula  = isinstance(raw_formula, str) and raw_formula.startswith('=')

                row_cells.append({
                    'coord':      cf.coordinate,
                    'formula':    formula_str,
                    'value':      value_str,
                    'is_formula': is_formula,
                })
            rows.append(row_cells)

        # 빈 행 제거 (모든 셀이 비어있는 행)
        rows = [r for r in rows if any(c['formula'] != '' for c in r)]

        sheets[name] = rows

    return {
        'sheet_names': wb_formula.sheetnames,
        'sheets': sheets,
    }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '파일을 선택해 주세요.'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx',):
        return jsonify({'error': f'"{ext}" 형식은 지원하지 않습니다. openpyxl은 .xlsx 파일만 지원합니다.'}), 400

    try:
        data = parse_workbook(f.read())
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': f'파일 파싱 오류: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
